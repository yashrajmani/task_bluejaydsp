import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Function to read employee data from an XLSX file
def read_data_from_xlsx(filename):
    employees = []

    # Load the Excel file
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Iterate through rows in the sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        position_id, _, time, time_out, timecard_hours, _, _, employee_name, _ = row

        # Skip rows with missing data
        if not (position_id and employee_name):
            continue

        # Check if time and time_out are not None before parsing
        if time and time_out:
            employees.append({
                "position_id": position_id,
                "employee_name": employee_name,
                "time": time,
                "time_out": time_out,
                "timecard_hours": timecard_hours
            })

    return employees



# Function to check if an employee has less than 10 hours between shifts but greater than 1 hour
def check_shift_hours(employees):
    print("Checking for employees who have less than 10 hours between shifts but greater than 1 hour...")

    for i in range(len(employees) - 1):
        current_employee = employees[i]
        next_employee = employees[i + 1]

        current_time_out = current_employee["time_out"]
        next_time = next_employee["time"]

        time_difference = (next_time - current_time_out).total_seconds() / 3600  # Convert to hours

        if 1 < time_difference < 10:
            print(f"Employee: {current_employee['employee_name']}, Position ID: {current_employee['position_id']} has less than 10 hours between shifts but greater than 1 hour.")

# Function to check if an employee has worked for 7 consecutive days
def check_consecutive_days(employees):
    print("Checking for employees who have worked for 7 consecutive days...")

    # Sort employees by time worked
    employees.sort(key=lambda x: x["time"])
    
    for i, employee in enumerate(employees[:-6]):  # Iterate up to the 6th last employee
        current_date = datetime.strptime(employee["time"], "%m/%d/%Y %I:%M %p").date()
        consecutive_days = 1

        for j in range(i + 1, len(employees)):
            next_date = datetime.strptime(employees[j]["time"], "%m/%d/%Y %I:%M %p").date()
            if (next_date - current_date).days == 1:
                consecutive_days += 1
                current_date = next_date
            else:
                break

            if consecutive_days == 7:
                print(f"Employee: {employee['employee_name']}, Position ID: {employee['position_id']} has worked for 7 consecutive days.\n")
                break



# Function to check if an employee has worked for 7 consecutive days
def check_consecutive_days(employees):
    print("Checking for employees who have worked for 7 consecutive days...")

    for employee in employees:
        consecutive_days = 1    
#FORMAT  09/13/2023 02:20 PM
        current_date = datetime.strptime(employee["time"], '%M/%D/%Y %H:%M %P').date()
        for next_employee in employees:
            next_date = datetime.strptime(next_employee["time"], '%M/%D/%Y %H:%M %P').date()
            if (next_date - current_date).days == 1:
                consecutive_days += 1
                current_date = next_date
            else:
                break
            if consecutive_days == 7:
                print(f"Employee: {employee['employee_name']}, Position ID: {employee['position_id']} has worked for 7 consecutive days.\n")
                break

# Function to check if an employee has worked for more than 14 hours in a single shift
def check_single_shift_hours(employees):
    print("Checking for employees who have worked for more than 14 hours in a single shift...")

    for employee in employees:
        timecard_hours = employee["timecard_hours"]
        if timecard_hours and int(timecard_hours.split(':')[0]) > 14:
            print(f"Employee: {employee['employee_name']}, Position ID: {employee['position_id']} has worked for more than 14 hours in a single shift.\n")

if __name__ == "__main__":
    filename = "Assignment_Timecard.xlsx"  
    employees = read_data_from_xlsx(filename)

    while True:
        print("Menu:")
        print("1. Check for employees who have worked for 7 consecutive days")
        print("2. Check for employees who have less than 10 hours between shifts but greater than 1 hour")
        print("3. Check for employees who have worked for more than 14 hours in a single shift")
        print("4. Exit")

        choice = input("Enter your choice: ")

        if choice == "1":
            check_consecutive_days(employees)
        elif choice == "2":
            check_shift_hours(employees)
        elif choice == "3":
            check_single_shift_hours(employees)
        elif choice == "4":
            print("Exiting program.")
            break
        else:
            print("Invalid choice. Please try again.")
